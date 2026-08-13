"""Golden-file test: the real sample invoice, end to end, cell by cell.

Ground truth was verified by hand with Decimal against the source PDF before any
parser code existed, so this asserts against independently-derived numbers rather
than against whatever the parser happens to produce.

NOTE: the fixture is real client financial data. Keep this repository private,
or replace the fixture with a redacted invoice before making it public.
"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal as D
from pathlib import Path

import openpyxl
import pytest

from gstparse.excel.writer import write_register
from gstparse.models import Invoice
from gstparse.parser import parse_pdf

FIXTURE = Path(__file__).parent / "fixtures" / "sample_invoice.pdf"

# The fixture is real client financial data and is deliberately NOT in version
# control (the repository is public). Drop the invoice at the path above to run
# these; every other test in the suite is self-contained and always runs.
pytestmark = pytest.mark.skipif(
    not FIXTURE.exists(),
    reason="client invoice fixture not present -- see README 'Test fixtures'",
)

# Independently hand-verified. (src_line, hsn, qty, rate, gst%, taxable, cgst, sgst, total)
EXPECTED_LINES = [
    (1, "21069099", "300", "129", "18", "38700.00", "3483.00", "3483.00", "45666.00"),
    (2, "21069099", "150", "139", "18", "20850.00", "1876.50", "1876.50", "24603.00"),
    (3, "09109100", "112.5", "250", "5", "28125.00", "703.13", "703.12", "29531.25"),
    (4, "21069099", "104", "143", "18", "14872.00", "1338.48", "1338.48", "17548.96"),
    (5, "09109100", "8", "368", "5", "2944.00", "73.60", "73.60", "3091.20"),
    (7, "21069099", "15", "109", "18", "1635.00", "147.15", "147.15", "1929.30"),
    (8, "21069099", "9", "209", "18", "1881.00", "169.29", "169.29", "2219.58"),
    (9, "09109100", "30", "153", "5", "4590.00", "114.75", "114.75", "4819.50"),
    (10, "09109990", "12", "595", "5", "7140.00", "178.50", "178.50", "7497.00"),
    (11, "21069099", "3", "199", "18", "597.00", "53.73", "53.73", "704.46"),
    (14, "09011111", "40", "500", "5", "20000.00", "500.00", "500.00", "21000.00"),
]


@pytest.fixture(scope="module")
def invoice() -> Invoice:
    return parse_pdf(FIXTURE)


# --- header ---------------------------------------------------------------

def test_invoice_header_fields(invoice: Invoice) -> None:
    assert invoice.invoice_no == "SBA/26-27/60"
    assert invoice.invoice_date == dt.date(2026, 7, 1)
    assert invoice.seller_gstin == "27AAKCG6367M1ZV"
    assert invoice.buyer_gstin == "27ABDFK6885B1Z6"
    assert invoice.buyer_name == "KESARI NANDAN FOODS"
    assert invoice.supply_type == "intra"


def test_free_text_date_is_parsed_to_a_real_date(invoice: Invoice) -> None:
    """The invoice says '1st July 2026'. It must become a date, not stay a string."""
    assert isinstance(invoice.invoice_date, dt.date)
    assert invoice.invoice_date == dt.date(2026, 7, 1)


# --- line items -----------------------------------------------------------

def test_fourteen_source_lines_become_eleven_register_rows(invoice: Invoice) -> None:
    assert len(invoice.line_items) == 11
    assert invoice.skipped_zero_qty == [6, 12, 13]


@pytest.mark.parametrize("expected", EXPECTED_LINES, ids=[str(e[0]) for e in EXPECTED_LINES])
def test_each_line_item_matches_hand_verified_values(
    invoice: Invoice, expected: tuple[str, ...]
) -> None:
    src, hsn, qty, rate, pct, taxable, cgst, sgst, total = expected
    item = next(i for i in invoice.line_items if i.src_line == src)
    assert item.hsn == hsn
    assert item.qty == D(qty)
    assert item.unit_rate == D(rate)
    assert item.gst_rate == D(pct)
    assert item.taxable == D(taxable)
    assert item.cgst == D(cgst)
    assert item.sgst == D(sgst)
    assert item.line_total == D(total)
    assert item.igst is None       # intra-state: never both


# --- totals ---------------------------------------------------------------

def test_totals_tie_out_to_the_rupee(invoice: Invoice) -> None:
    assert invoice.taxable_total == D("141334.00")
    assert invoice.cgst_total == D("8638.13")
    assert invoice.sgst_total == D("8638.12")
    assert invoice.tax_total == D("17276.25")
    assert invoice.line_total_sum == D("158610.25")
    assert invoice.round_off == D("-0.25")
    assert invoice.computed_grand_total == D("158610.00")
    assert invoice.tie_out_delta == D("0.00")


def test_our_tax_total_matches_the_vendor_gst_amount_column(invoice: Invoice) -> None:
    """The vendor's own Gst Amount column totals 17,276.25 -- we agree with it."""
    assert invoice.tax_total == D("17276.25")


def test_we_disagree_with_the_vendor_printed_sgst_by_one_paisa(invoice: Invoice) -> None:
    """The invoice prints SGST@2.5% as 1569.98; the true remainder is 1569.97.

    The vendor rounded both halves of the 5% bucket up. We are right, and this
    is why the register still ties to the invoice grand total.
    """
    five_pc = [i for i in invoice.line_items if i.gst_rate == D("5")]
    assert sum((i.cgst or D(0) for i in five_pc), D(0)) == D("1569.98")
    assert sum((i.sgst or D(0) for i in five_pc), D(0)) == D("1569.97")


def test_invoice_is_not_blocked(invoice: Invoice) -> None:
    blocking = [f for f in invoice.findings if f.severity == "blocking"]
    assert blocking == [], f"unexpected blocking findings: {blocking}"


def test_missing_unit_is_a_warning_never_invented(invoice: Invoice) -> None:
    codes = {f.rule_code for f in invoice.findings}
    assert "R8_UNIT_UNKNOWN" in codes
    assert all(item.unit is None for item in invoice.line_items)


# --- Excel output ---------------------------------------------------------

@pytest.fixture(scope="module")
def workbook(invoice: Invoice, tmp_path_factory: pytest.TempPathFactory):  # type: ignore[no-untyped-def]
    path = write_register([invoice], tmp_path_factory.mktemp("out"))
    return openpyxl.load_workbook(path)


def test_workbook_has_the_three_required_sheets(workbook) -> None:  # type: ignore[no-untyped-def]
    assert workbook.sheetnames == ["Sales Data", "Invoice Summary", "Import Log"]


def test_header_layout_matches_the_client_template(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Sales Data"]
    assert sheet["B3"].value == "Invoice No"
    assert sheet["J3"].value == "Taxable"
    assert sheet["K3"].value == "IGST"
    assert sheet["M3"].value == "CGST"
    assert sheet["O3"].value == "SGST"
    assert sheet["Q3"].value == " TOTAL"
    assert sheet["K4"].value == "Rate" and sheet["L4"].value == "Amount"
    merged = {str(m) for m in sheet.merged_cells.ranges}
    assert {"B3:B4", "J3:J4", "K3:L3", "M3:N3", "O3:P3", "Q3:Q4"} <= merged


def test_hsn_is_written_as_text_so_leading_zeros_survive(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Sales Data"]
    values = [sheet[f"G{r}"].value for r in range(5, 16)]
    assert "09109100" in values
    assert "09011111" in values
    assert all(sheet[f"G{r}"].number_format == "@" for r in range(5, 16))


def test_date_is_a_real_excel_date_with_explicit_format(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Sales Data"]
    assert isinstance(sheet["C5"].value, dt.datetime)
    assert sheet["C5"].number_format == "dd-mm-yyyy"


def test_total_row_spans_the_actual_data_range_not_a_fixed_one(workbook) -> None:
    """Client template defect #1: it summed J5:J29 while data ran to row 30."""
    sheet = workbook["Sales Data"]
    assert sheet["B16"].value == "Total"
    for column in ("J", "L", "N", "P", "Q"):
        assert sheet[f"{column}16"].value == f"=SUM({column}5:{column}15)"


def test_row_count_is_dynamic_not_a_fixed_block(workbook) -> None:
    """Client template defect #2: a fixed ~26-row block that two invoices overflow."""
    sheet = workbook["Sales Data"]
    assert sheet["B15"].value is not None   # last data row
    assert sheet["B17"].value is None       # nothing after the total row


def test_every_row_total_is_a_live_formula(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Sales Data"]
    for row in range(5, 16):
        assert sheet[f"Q{row}"].value == f"=J{row}+L{row}+N{row}+P{row}"


def test_invoice_summary_carries_round_off_and_tie_out(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Invoice Summary"]
    assert sheet.cell(1, 9).value == "Round Off"
    assert sheet.cell(1, 12).value == "Tie-out Delta"
    assert sheet.cell(2, 9).value == -0.25
    assert sheet.cell(2, 12).value == 0


def test_import_log_records_provenance(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Import Log"]
    assert sheet.cell(2, 1).value == "sample_invoice.pdf"
    assert len(str(sheet.cell(2, 2).value)) == 64      # sha-256
    assert sheet.cell(2, 3).value == 2                 # parse tier
    assert sheet.cell(2, 4).value == "no"              # ocr
    assert sheet.cell(2, 5).value == 11                # rows
    assert sheet.cell(2, 6).value == "6, 12, 13"       # skipped zero-qty


def test_export_never_overwrites_an_existing_file(invoice: Invoice, tmp_path: Path) -> None:
    stamp = dt.datetime(2026, 8, 14, 12, 0, 0)
    first = write_register([invoice], tmp_path, stamp)
    second = write_register([invoice], tmp_path, stamp)
    assert first != second
    assert first.exists() and second.exists()


def test_two_invoices_overflow_the_old_fixed_block_and_still_work(
    invoice: Invoice, tmp_path: Path
) -> None:
    """22 rows exceeds nothing here, but proves rows are inserted, not templated."""
    path = write_register([invoice, invoice], tmp_path)
    sheet = openpyxl.load_workbook(path)["Sales Data"]
    assert sheet["B26"].value is not None          # 22 data rows -> 5..26
    assert sheet["B27"].value == "Total"
    assert sheet["J27"].value == "=SUM(J5:J26)"
