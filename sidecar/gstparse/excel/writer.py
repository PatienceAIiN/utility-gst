"""Sales Register writer (brief §7).

Built rather than copied. The client template has a fixed ~26-row data block and
a total row that sums J5:J29 while data runs to row 30 -- so the last row is
silently excluded. Copying it would inherit both defects, so the layout is
reconstructed and the totals always span the ACTUAL data range.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Sequence
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Invoice

HEADER_ROW_TOP: Final = 3
HEADER_ROW_SUB: Final = 4
DATA_START: Final = 5

MONEY_FMT: Final = "#,##0.00"
QTY_FMT: Final = "#,##0.###"      # brief §5: quantity can be fractional (112.5)
DATE_FMT: Final = "dd-mm-yyyy"    # brief §5: a real Excel date, never a string
TEXT_FMT: Final = "@"             # brief §5: HSN must not be coerced to a number
RATE_FMT: Final = '0.00"%"'

# Column letter -> (top header, sub header, width, number format)
COLUMNS: Final[tuple[tuple[str, str | None, str | None, float, str], ...]] = (
    ("B", "Invoice No", None, 15.0, TEXT_FMT),
    ("C", "Date", None, 12.0, DATE_FMT),
    ("D", "Party Name", None, 28.0, TEXT_FMT),
    ("E", "GST No", None, 20.0, TEXT_FMT),
    ("F", "Product Description", None, 30.0, TEXT_FMT),
    ("G", "HSN Code", None, 13.0, TEXT_FMT),
    ("H", "QTY", None, 9.0, QTY_FMT),
    ("I", "Unit", None, 8.0, TEXT_FMT),
    ("J", "Taxable", None, 14.0, MONEY_FMT),
    ("K", "IGST", "Rate", 8.0, RATE_FMT),
    ("L", "IGST", "Amount", 13.0, MONEY_FMT),
    ("M", "CGST", "Rate", 8.0, RATE_FMT),
    ("N", "CGST", "Amount", 13.0, MONEY_FMT),
    ("O", "SGST", "Rate", 8.0, RATE_FMT),
    ("P", "SGST", "Amount", 13.0, MONEY_FMT),
    ("Q", " TOTAL", None, 15.0, MONEY_FMT),
)

_THIN = Side(style="thin", color="FF000000")
BORDER: Final = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEAD_FILL: Final = PatternFill("solid", fgColor="FFC6E0B4")
HEAD_FONT: Final = Font(bold=True)
CENTRE: Final = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _num(value: Decimal | None) -> float | None:
    """openpyxl cannot store Decimal. Convert only at the boundary, never before.

    All arithmetic upstream is exact Decimal; this is the final render step and
    the value is already quantised to 2 dp, so the float is display-only.
    """
    return None if value is None else float(value)


def _write_header(sheet: Worksheet) -> None:
    for letter, top, sub, width, _fmt in COLUMNS:
        sheet.column_dimensions[letter].width = width
        top_cell = sheet[f"{letter}{HEADER_ROW_TOP}"]
        top_cell.value = top
        top_cell.font, top_cell.fill = HEAD_FONT, HEAD_FILL
        top_cell.alignment, top_cell.border = CENTRE, BORDER

        sub_cell = sheet[f"{letter}{HEADER_ROW_SUB}"]
        sub_cell.value = sub
        sub_cell.font, sub_cell.fill = HEAD_FONT, HEAD_FILL
        sub_cell.alignment, sub_cell.border = CENTRE, BORDER

    # Vertical merges for single-level columns; horizontal pairs for the tax groups.
    for letter, _top, sub, _w, _f in COLUMNS:
        if sub is None:
            sheet.merge_cells(f"{letter}{HEADER_ROW_TOP}:{letter}{HEADER_ROW_SUB}")
    for left, right in (("K", "L"), ("M", "N"), ("O", "P")):
        sheet.merge_cells(f"{left}{HEADER_ROW_TOP}:{right}{HEADER_ROW_TOP}")


def _write_row(sheet: Worksheet, row: int, values: dict[str, object]) -> None:
    for letter, _top, _sub, _w, fmt in COLUMNS:
        cell = sheet[f"{letter}{row}"]
        cell.value = values.get(letter)
        cell.number_format = fmt
        cell.border = BORDER
    # Brief §7: TOTAL is a live formula, not a baked number.
    total = sheet[f"Q{row}"]
    total.value = f"=J{row}+L{row}+N{row}+P{row}"
    total.number_format = MONEY_FMT
    total.border = BORDER


def _write_totals(sheet: Worksheet, first: int, last: int) -> int:
    """Total row spanning the ACTUAL data range -- the template's defect #1."""
    row = last + 1
    label = sheet[f"B{row}"]
    label.value = "Total"
    label.font = HEAD_FONT
    sheet.merge_cells(f"B{row}:I{row}")
    for letter in ("B", "C", "D", "E", "F", "G", "H", "I"):
        sheet[f"{letter}{row}"].border = BORDER
    for letter in ("J", "K", "L", "M", "N", "O", "P", "Q"):
        cell = sheet[f"{letter}{row}"]
        if letter in ("K", "M", "O"):       # rate columns do not sum meaningfully
            cell.value = None
        else:
            cell.value = f"=SUM({letter}{first}:{letter}{last})"
        cell.number_format = MONEY_FMT
        cell.font, cell.border = HEAD_FONT, BORDER
    return row


def build_sales_data(sheet: Worksheet, invoices: Sequence[Invoice]) -> tuple[int, int]:
    _write_header(sheet)
    row = DATA_START
    for invoice in invoices:
        for item in invoice.line_items:
            _write_row(sheet, row, {
                "B": invoice.invoice_no,
                "C": invoice.invoice_date,
                "D": invoice.buyer_name,
                "E": invoice.buyer_gstin,
                "F": item.description,
                "G": item.hsn,          # text format keeps the leading zero
                "H": _num(item.qty),
                "I": item.unit,
                "J": _num(item.taxable),
                "K": _num(item.igst_rate),
                "L": _num(item.igst),
                "M": _num(item.cgst_rate),
                "N": _num(item.cgst),
                "O": _num(item.sgst_rate),
                "P": _num(item.sgst),
            })
            row += 1
    last = row - 1
    if last >= DATA_START:
        _write_totals(sheet, DATA_START, last)
    sheet.freeze_panes = f"A{DATA_START}"
    return DATA_START, last


def build_invoice_summary(sheet: Worksheet, invoices: Sequence[Invoice]) -> None:
    """The reconciliation artifact (brief §7).

    Round-off is invoice-level and has nowhere to live on Sales Data, so this is
    the only sheet where the register can be tied back to the invoice totals
    (DECISIONS.md R-13).
    """
    headers = ("Invoice No", "Date", "Party Name", "GSTIN", "Taxable", "IGST", "CGST",
               "SGST", "Round Off", "Computed Total", "Invoice Total", "Tie-out Delta")
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font, cell.fill, cell.border, cell.alignment = HEAD_FONT, HEAD_FILL, BORDER, CENTRE
    widths = (15, 12, 26, 20, 14, 12, 12, 12, 11, 15, 14, 13)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for offset, invoice in enumerate(invoices, start=2):
        values = (invoice.invoice_no, invoice.invoice_date, invoice.buyer_name,
                  invoice.buyer_gstin, _num(invoice.taxable_total), _num(invoice.igst_total),
                  _num(invoice.cgst_total), _num(invoice.sgst_total), _num(invoice.round_off),
                  _num(invoice.computed_grand_total), _num(invoice.stated_grand_total),
                  _num(invoice.tie_out_delta))
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=index, value=value)
            cell.border = BORDER
            if index == 2:
                cell.number_format = DATE_FMT
            elif index >= 5:
                cell.number_format = MONEY_FMT
    sheet.freeze_panes = "A2"


def build_import_log(sheet: Worksheet, invoices: Sequence[Invoice], when: dt.datetime) -> None:
    headers = ("Source File", "SHA-256", "Parse Tier", "OCR", "Rows Produced",
               "Skipped (zero qty)", "Warnings", "Blocking", "Imported At")
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font, cell.fill, cell.border, cell.alignment = HEAD_FONT, HEAD_FILL, BORDER, CENTRE
    for index, width in enumerate((34, 66, 10, 7, 14, 18, 40, 10, 20), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for offset, invoice in enumerate(invoices, start=2):
        # Info-level notes (a healed column, merged continuation pages) are kept
        # in the log too: they record what the parser did to the source and an
        # audit needs that, even though they are not shown as warnings in the app.
        warnings = "; ".join(
            f.rule_code for f in invoice.findings if f.severity in ("warning", "info"))
        blocking = "; ".join(f.rule_code for f in invoice.findings if f.severity == "blocking")
        values: tuple[str | int, ...] = (
            invoice.source_file, invoice.sha256, invoice.parse_tier,
            "yes" if invoice.ocr_used else "no", len(invoice.line_items),
            ", ".join(str(n) for n in invoice.skipped_zero_qty),
            warnings, blocking or "none", when.strftime("%Y-%m-%d %H:%M:%S"))
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=index, value=value)
            cell.border = BORDER
            if index == 2:
                cell.number_format = TEXT_FMT
    sheet.freeze_panes = "A2"


def write_register(
    invoices: Sequence[Invoice], out_dir: Path, when: dt.datetime | None = None
) -> Path:
    """Write a NEW timestamped register. Never overwrites an existing export (§7)."""
    stamp = when or dt.datetime.now()  # noqa: DTZ005 - export filenames use local wall-clock
    workbook = Workbook()
    sales = workbook.active
    assert sales is not None
    sales.title = "Sales Data"
    build_sales_data(sales, invoices)
    build_invoice_summary(workbook.create_sheet("Invoice Summary"), invoices)
    build_import_log(workbook.create_sheet("Import Log"), invoices, stamp)

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"Sales-Register-{stamp:%Y%m%d-%H%M%S}.xlsx"
    counter = 1
    while path.exists():
        path = out_dir / f"Sales-Register-{stamp:%Y%m%d-%H%M%S}-{counter}.xlsx"
        counter += 1
    workbook.save(path)
    return path
