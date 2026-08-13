"""In-app spreadsheet viewing and editing for CSV and XLSX.

Format handling is Microsoft-compatible using open-source implementations only:

* **XLSX** via ``openpyxl`` -- a direct implementation of Office Open XML
  (ECMA-376 / ISO-IEC 29500), the same format Excel itself reads and writes.
  No Excel, no COM, no proprietary runtime.
* **CSV** via the stdlib ``csv`` module driven with Excel's own dialect:
  ``QUOTE_MINIMAL``, ``"`` quoting with doubling, and CRLF terminators. Files
  are written UTF-8 **with BOM**, because that is the only encoding signal
  Excel reliably honours -- a plain UTF-8 CSV opens as mojibake for anything
  non-ASCII, which for Indian party names is most of them.

The delimiter detected on read is preserved on write, so a semicolon-separated
export (what Excel produces under a locale whose list separator is ``;``) does
not silently become comma-separated when saved.

Writes are non-destructive by default: an operator may open a source document
for an already-imported invoice, so saving produces a timestamped copy unless
overwrite is explicitly requested.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

MAX_ROWS: Final = 20_000      # a grid past this is unusable in the UI anyway
MAX_COLS: Final = 200
CSV_SNIFF_BYTES: Final = 64 * 1024

# Excel writes CRLF and reads UTF-8 only when a BOM is present.
EXCEL_ENCODING: Final = "utf-8-sig"
EXCEL_LINETERM: Final = "\r\n"


def _detect_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def read_csv(path: Path) -> dict[str, Any]:
    # utf-8-sig strips a BOM if present and is a no-op otherwise.
    raw = path.read_text(encoding=EXCEL_ENCODING, errors="replace")
    delimiter = _detect_delimiter(raw[:CSV_SNIFF_BYTES])
    rows = list(csv.reader(io.StringIO(raw), delimiter=delimiter, quotechar='"'))
    truncated = len(rows) > MAX_ROWS
    rows = rows[:MAX_ROWS]
    width = min(max((len(r) for r in rows), default=0), MAX_COLS)
    normalised = [[(r[i] if i < len(r) else "") for i in range(width)] for r in rows]
    return {
        "kind": "csv",
        "sheets": [{"name": path.stem, "rows": normalised}],
        "active": 0,
        "truncated": truncated,
        "delimiter": delimiter,
        "path": str(path),
    }


def read_xlsx(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    truncated = False
    for worksheet in workbook.worksheets:
        rows: list[list[str]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS:
                truncated = True
                break
            rows.append(["" if cell is None else str(cell) for cell in row[:MAX_COLS]])
        width = min(max((len(r) for r in rows), default=0), MAX_COLS)
        rows = [[(r[i] if i < len(r) else "") for i in range(width)] for r in rows]
        sheets.append({"name": worksheet.title, "rows": rows})
    workbook.close()
    return {"kind": "xlsx", "sheets": sheets, "active": 0,
            "truncated": truncated, "delimiter": ",", "path": str(path)}


def read_sheet(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    raise ValueError(f"Cannot open {suffix} as a spreadsheet")


def _target(path: Path, overwrite: bool) -> Path:
    """Never clobber a source document unless explicitly told to."""
    if overwrite:
        return path
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")  # noqa: DTZ005 - local wall-clock
    return path.with_name(f"{path.stem}-edited-{stamp}{path.suffix}")


def write_sheet(
    path: Path,
    sheets: list[dict[str, Any]],
    overwrite: bool = False,
    delimiter: str = ",",
) -> dict[str, Any]:
    destination = _target(path, overwrite)
    suffix = destination.suffix.lower()

    if suffix == ".csv":
        rows = sheets[0]["rows"] if sheets else []
        with open(destination, "w", newline="", encoding=EXCEL_ENCODING) as handle:
            csv.writer(
                handle,
                delimiter=delimiter or ",",
                quotechar='"',
                quoting=csv.QUOTE_MINIMAL,
                lineterminator=EXCEL_LINETERM,
            ).writerows(rows)
    elif suffix in (".xlsx", ".xlsm"):
        workbook = Workbook()
        default = workbook.active
        if default is not None:
            workbook.remove(default)
        for sheet in sheets or [{"name": "Sheet1", "rows": []}]:
            worksheet = workbook.create_sheet(title=str(sheet.get("name") or "Sheet1")[:31])
            rows = sheet.get("rows", [])
            for row in rows:
                # Everything round-trips as TEXT. Letting Excel re-infer types
                # would destroy leading zeros on HSN codes and reinterpret
                # dd-mm-yyyy as US dates -- both silent, both unacceptable here.
                worksheet.append([("" if c is None else str(c)) for c in row])
            for cell_row in worksheet.iter_rows():
                for cell in cell_row:
                    cell.number_format = "@"
            width = max((len(r) for r in rows), default=0)
            for index in range(1, min(width, MAX_COLS) + 1):
                worksheet.column_dimensions[get_column_letter(index)].width = 18
        workbook.save(destination)
    else:
        raise ValueError(f"Cannot save {suffix}")

    return {"path": str(destination), "overwrote": overwrite}
